from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Any
from xml.etree.ElementTree import ParseError
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet


class XlsxImportParserError(ValueError):
    """A stored XLSX file cannot be represented as lossless ImportRow raw data."""


@dataclass(frozen=True)
class ParsedXlsxRow:
    row_number: int
    sheet_name: str
    raw_data_json: dict[str, Any]


def parse_xlsx_file(path: Path) -> list[ParsedXlsxRow]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except OSError as exc:
        raise XlsxImportParserError("Stored XLSX file cannot be read.") from exc
    except (BadZipFile, InvalidFileException, KeyError, ParseError, ValueError) as exc:
        raise XlsxImportParserError("XLSX file is invalid.") from exc

    try:
        worksheet = next((sheet for sheet in workbook.worksheets if sheet.sheet_state == "visible"), None)
        if worksheet is None:
            raise XlsxImportParserError("XLSX workbook has no visible worksheet.")
        return _parse_worksheet_rows(worksheet)
    finally:
        workbook.close()


def _parse_worksheet_rows(worksheet: Worksheet) -> list[ParsedXlsxRow]:
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows or not any(_has_value(value) for row in rows for value in row):
        raise XlsxImportParserError("XLSX worksheet is empty.")

    header_values = list(rows[0])
    if not any(_has_value(value) for value in header_values):
        raise XlsxImportParserError("XLSX header is missing.")

    last_header_column = max(index for index, value in enumerate(header_values) if _has_value(value)) + 1
    headers = [_serialize_header(value) for value in header_values[:last_header_column]]
    if any(not header.strip() for header in headers):
        raise XlsxImportParserError("XLSX header contains an empty column name.")
    if len(headers) != len(set(headers)):
        raise XlsxImportParserError("XLSX header contains duplicate column names.")

    parsed_rows: list[ParsedXlsxRow] = []
    for row_number, source_values in enumerate(rows[1:], start=2):
        values = list(source_values)
        if any(_has_value(value) for value in values[len(headers) :]):
            raise XlsxImportParserError("XLSX row has more values than the header.")

        padded_values = [*values[: len(headers)], *([None] * (len(headers) - len(values)))]
        serialized_values = [_serialize_cell(value) for value in padded_values]
        if all(value == "" for value in serialized_values):
            continue
        parsed_rows.append(
            ParsedXlsxRow(
                row_number=row_number,
                sheet_name=worksheet.title,
                raw_data_json=dict(zip(headers, serialized_values, strict=True)),
            )
        )

    return parsed_rows


def _has_value(value: object) -> bool:
    return value is not None and value != ""


def _serialize_header(value: object) -> str:
    serialized = _serialize_cell(value)
    return str(serialized)


def _serialize_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return value
